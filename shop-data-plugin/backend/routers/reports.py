"""
报表生成路由
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime, timedelta
from decimal import Decimal
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from database import get_db
from models.user import User
from models.order import Order
from utils.auth import get_current_user

router = APIRouter(prefix="/api/reports", tags=["报表导出"])


@router.get("/orders/excel", summary="导出订单报表Excel")
async def export_orders_excel(
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    status: Optional[str] = Query(None, description="订单状态"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出订单报表为Excel格式"""
    query = db.query(Order).filter(Order.user_id == current_user.id)

    # 筛选条件
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Order.order_time >= start_dt)
        except:
            pass

    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Order.order_time < end_dt)
        except:
            pass

    if status:
        query = query.filter(Order.status == status)

    orders = query.order_by(Order.order_time.desc()).all()

    # 创建DataFrame
    data = []
    for order in orders:
        data.append({
            "订单号": order.order_id,
            "平台": order.platform,
            "商品名称": order.product_name,
            "规格": order.sku_name,
            "数量": order.quantity,
            "销售单价": float(order.sale_price),
            "订单金额": float(order.total_amount),
            "成本单价": float(order.cost_price),
            "运费": float(order.shipping_cost),
            "佣金": float(order.commission),
            "广告费": float(order.ad_cost),
            "其他费用": float(order.other_cost),
            "毛利润": float(order.gross_profit),
            "净利润": float(order.net_profit),
            "订单状态": order.status,
            "下单时间": order.order_time.strftime("%Y-%m-%d %H:%M:%S") if order.order_time else "",
            "买家姓名": order.buyer_name,
            "省份": order.province,
            "城市": order.city,
            "区县": order.district
        })

    df = pd.DataFrame(data)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "订单报表"

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            else:
                cell.alignment = Alignment(horizontal="center")

    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 添加汇总行
    if data:
        summary_row = len(data) + 2
        ws.cell(row=summary_row, column=1, value="汇总")
        ws.cell(row=summary_row, column=5, value=sum(item["数量"] for item in data))
        ws.cell(row=summary_row, column=7, value=sum(item["订单金额"] for item in data))
        ws.cell(row=summary_row, column=14, value=sum(item["毛利润"] for item in data))
        ws.cell(row=summary_row, column=15, value=sum(item["净利润"] for item in data))

        # 汇总行样式
        for c in range(1, len(data[0]) + 1):
            cell = ws.cell(row=summary_row, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    filename = f"订单报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/profit/excel", summary="导出利润分析报表")
async def export_profit_excel(
    start_date: Optional[str] = Query(None, description="开始日期"),
    end_date: Optional[str] = Query(None, description="结束日期"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出利润分析报表"""
    query = db.query(Order).filter(Order.user_id == current_user.id)

    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Order.order_time >= start_dt)
        except:
            pass

    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Order.order_time < end_dt)
        except:
            pass

    orders = query.all()

    # 计算汇总数据
    total_orders = len(orders)
    total_quantity = sum(o.quantity for o in orders)
    total_sales = sum(o.total_amount for o in orders)
    total_product_cost = sum(o.cost_price * o.quantity for o in orders)
    total_shipping = sum(o.shipping_cost for o in orders)
    total_commission = sum(o.commission for o in orders)
    total_ad_cost = sum(o.ad_cost for o in orders)
    total_other = sum(o.other_cost for o in orders)
    total_cost = total_product_cost + total_shipping + total_commission + total_ad_cost + total_other
    total_profit = sum(o.gross_profit for o in orders)
    avg_profit_rate = (total_profit / total_sales * 100) if total_sales > 0 else 0

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "利润分析"

    # 样式
    title_font = Font(bold=True, size=16)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # 标题
    ws.merge_cells('A1:D1')
    ws['A1'] = "利润分析报告"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")

    # 报告信息
    ws['A3'] = "报告生成时间："
    ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A4'] = "统计周期："
    ws['B4'] = f"{start_date or '起始'} 至 {end_date or '至今'}"

    # 汇总表
    ws['A6'] = "销售汇总"
    ws['A6'].font = Font(bold=True, size=12)

    summary_data = [
        ["指标", "数值"],
        ["总订单数", total_orders],
        ["总销售数量", total_quantity],
        ["总销售额", f"¥{total_sales:.2f}"],
        ["平均客单价", f"¥{(total_sales / total_orders if total_orders > 0 else 0):.2f}"],
        ["", ""],
        ["成本构成", ""],
        ["商品成本", f"¥{total_product_cost:.2f}"],
        ["运费成本", f"¥{total_shipping:.2f}"],
        ["平台佣金", f"¥{total_commission:.2f}"],
        ["广告费用", f"¥{total_ad_cost:.2f}"],
        ["其他费用", f"¥{total_other:.2f}"],
        ["总成本", f"¥{total_cost:.2f}"],
        ["", ""],
        ["利润分析", ""],
        ["毛利润", f"¥{total_profit:.2f}"],
        ["毛利率", f"{avg_profit_rate:.2f}%"]
    ]

    for r_idx, row in enumerate(summary_data, 7):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 7:
                cell.fill = header_fill
                cell.font = header_font

    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"利润分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/daily/excel", summary="导出每日统计报表")
async def export_daily_excel(
    start_date: Optional[str] = Query(None, description="开始日期"),
    end_date: Optional[str] = Query(None, description="结束日期"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出每日统计报表"""
    from sqlalchemy import func

    query = db.query(
        func.date(Order.order_time).label("date"),
        func.count(Order.id).label("orders"),
        func.sum(Order.quantity).label("quantity"),
        func.sum(Order.total_amount).label("sales"),
        func.sum(Order.gross_profit).label("profit")
    ).filter(Order.user_id == current_user.id)

    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Order.order_time >= start_dt)
        except:
            pass

    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Order.order_time < end_dt)
        except:
            pass

    results = query.group_by(func.date(Order.order_time)).order_by(func.date(Order.order_time).desc()).all()

    # 创建DataFrame
    data = []
    for r in results:
        profit_rate = (r.profit / r.sales * 100) if r.sales and r.sales > 0 else 0
        data.append({
            "日期": str(r.date),
            "订单数": r.orders,
            "销售数量": r.quantity,
            "销售额": float(r.sales or 0),
            "利润": float(r.profit or 0),
            "利润率": f"{profit_rate:.2f}%"
        })

    df = pd.DataFrame(data)

    # 创建Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="每日统计", index=False)
    output.seek(0)

    filename = f"每日统计报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )
